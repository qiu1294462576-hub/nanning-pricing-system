#!/usr/bin/env python3
"""
南宁家政成本核算与保本定价模块 — Phase 3
============================================
核心功能：
1. 四级成本分摊模型（直接成本→间接成本→风险成本→利润预留）
2. 全场景保本价测算
3. 历史订单盈利复盘
4. 导出保本定价测算表

用法:
    from core.cost_calculator import CostCalculator
    cc = CostCalculator()
    result = cc.calculate_cost("日常保洁", "熟练", "青秀", 3)
    print(result.break_even_price)
"""

import os
import csv
import math
from dataclasses import dataclass, field, asdict
from typing import Optional, List, Dict, Tuple
from datetime import datetime

# ============================================================
# Data Structures
# ============================================================

@dataclass
class CostDetail:
    """单次服务的完整成本明细"""
    service_type: str          # 服务类型
    skill_level: str           # 师傅等级
    district: str              # 城区
    duration: float            # 服务时长（小时）
    apartment_type: str        # 户型（影响耗材估算）
    time_period: str           # 时间段

    # 成本构成
    labor_cost: float = 0.0          # 人力成本（含社保）
    material_cost: float = 0.0       # 耗材成本
    transport_cost: float = 0.0      # 交通成本
    management_cost: float = 0.0     # 管理成本
    risk_buffer: float = 0.0         # 风险裕量

    # 汇总
    total_direct_cost: float = 0.0   # 直接成本小计
    total_cost: float = 0.0          # 总成本（含管理费）
    break_even_price: float = 0.0    # 保本价
    target_price_25: float = 0.0     # 目标售价（25%毛利率）
    target_price_35: float = 0.0     # 目标售价（35%毛利率）

    def to_dict(self) -> dict:
        return {k: round(v, 2) if isinstance(v, float) else v
                for k, v in asdict(self).items()}

    def summary_dict(self) -> dict:
        """简化的摘要字典，用于 Excel 输出"""
        return {
            "服务类型": self.service_type,
            "师傅等级": self.skill_level,
            "城区": self.district,
            "服务时长(h)": self.duration,
            "户型": self.apartment_type,
            "时间段": self.time_period,
            "人力成本": round(self.labor_cost, 2),
            "耗材成本": round(self.material_cost, 2),
            "交通成本": round(self.transport_cost, 2),
            "管理成本": round(self.management_cost, 2),
            "风险裕量": round(self.risk_buffer, 2),
            "总成本": round(self.total_cost, 2),
            "保本价": round(self.break_even_price, 2),
            "目标售价(25%毛利)": round(self.target_price_25, 2),
            "目标售价(35%毛利)": round(self.target_price_35, 2),
        }


# ============================================================
# Main Calculator
# ============================================================

class CostCalculator:
    """成本核算与保本定价计算器"""

    def __init__(self, params_path: Optional[str] = None):
        """
        初始化计算器，加载成本参数。

        Args:
            params_path: cost_parameters.csv 的路径，默认使用项目路径
        """
        if params_path is None:
            base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            params_path = os.path.join(base, "data", "nanning", "cost_parameters.csv")

        if not os.path.exists(params_path):
            raise FileNotFoundError(f"成本参数文件未找到: {params_path}")

        self.params_path = params_path
        self.params: Dict[str, Dict[str, float]] = self._load_params()

        # 服务类型列表
        self.service_types = ["日常保洁", "深度清洁", "开荒保洁", "家电清洗"]
        self.skill_levels = ["新手", "熟练", "金牌"]
        self.districts = ["青秀", "良庆", "江南", "西乡塘", "兴宁", "邕宁", "武鸣"]
        self.apartment_types = ["1室", "2室", "3室", "4室+", "别墅"]
        self.time_periods = ["工作日", "周末", "节假日", "旺季"]

        # 户型对耗材的影响系数（户型越大耗材越多）
        self.apartment_material_coeff = {
            "1室": 0.6, "2室": 0.8, "3室": 1.0, "4室+": 1.3, "别墅": 1.8,
        }

    # ----------------------------------------------------------
    # 参数加载
    # ----------------------------------------------------------

    def _load_params(self) -> Dict[str, Dict[str, float]]:
        """从 CSV 加载成本参数，并进行有效性校验"""
        params = {}
        errors = []
        with open(self.params_path, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for line, row in enumerate(reader, 2):
                cat = row["category"].strip()
                sub = row["subcategory"].strip()
                raw_val = row["value"].strip()
                try:
                    val = float(raw_val)
                except (ValueError, TypeError):
                    errors.append(f"第 {line} 行: 数值无效 '{raw_val}'")
                    continue
                # 基本合理性校验
                if cat in ("社保系数", "管理费率", "风险裕量", "目标毛利率"):
                    if val <= 0 or val >= 2:
                        errors.append(f"第 {line} 行: {cat} 值 {val} 超出合理范围 [0, 2]")
                if cat == "师傅到手时薪" and (val < 10 or val > 100):
                    errors.append(f"第 {line} 行: 师傅时薪 {val} 元/小时超出合理范围 [10, 100]")
                if cat not in params:
                    params[cat] = {}
                params[cat][sub] = val

        if errors:
            print("[警告] 参数校验发现以下问题:")
            for e in errors:
                print(f"  {e}")
        return params

    def get_param(self, category: str, subcategory: str = "通用", default: float = 0.0) -> float:
        """获取参数值"""
        return self.params.get(category, {}).get(subcategory, default)

    # ----------------------------------------------------------
    # 核心计算
    # ----------------------------------------------------------

    def calculate_cost(
        self,
        service_type: str,
        skill_level: str,
        district: str,
        duration: float,
        apartment_type: str = "3室",
        time_period: str = "工作日",
    ) -> CostDetail:
        """
        计算单次服务的完整成本。

        四级分摊模型：
        1. 直接成本：人力（师傅时薪×社保系数×时长） + 耗材（按服务类型×户型系数）
        2. 间接成本：交通（按城区距离） + 管理费（订单金额×6%）
        3. 风险成本：直接成本×10%（工时偏差补偿）
        4. 利润预留：25%-35%毛利率

        Returns:
            CostDetail 对象，包含完整的成本分解
        """
        detail = CostDetail(
            service_type=service_type,
            skill_level=skill_level,
            district=district,
            duration=duration,
            apartment_type=apartment_type,
            time_period=time_period,
        )

        # 校验服务时长
        if duration <= 0:
            raise ValueError(f"服务时长必须大于 0，当前值: {duration}")

        # ---- 1. 人力成本（受时间段影响） ----
        hourly_wage = self.get_param("师傅到手时薪", skill_level, 0)
        social_insurance = self.get_param("社保系数", "通用", 1.0)
        time_mult = self.get_param("时间段加价", time_period, 1.0)
        detail.labor_cost = hourly_wage * social_insurance * duration * time_mult

        # ---- 2. 耗材成本 ----
        base_material = self.get_param("耗材费率", service_type, 0)
        apt_coeff = self.apartment_material_coeff.get(apartment_type, 1.0)

        if service_type == "开荒保洁":
            # 开荒保洁按面积：使用生产率估算面积
            productivity = self.get_param("开荒保洁生产率", skill_level, 15)
            area = duration * productivity  # ㎡
            detail.material_cost = base_material * area
            detail._estimated_area = area  # 内部记录，不输出到 CostDetail
        elif service_type == "家电清洗":
            # 家电清洗按台数：估算台数
            hours_per_unit = self.get_param("家电清洗台数折算", "通用", 1.5)
            units = max(1, round(duration / hours_per_unit))
            detail.material_cost = base_material * units
            detail._estimated_units = units
        else:
            # 日常保洁/深度清洁按时计费
            detail.material_cost = base_material * apt_coeff

        # ---- 3. 交通成本 ----
        detail.transport_cost = self.get_param("交通加价", district, 0)

        # ---- 直接成本小计 ----
        detail.total_direct_cost = detail.labor_cost + detail.material_cost + detail.transport_cost

        # ---- 4. 风险裕量（直接成本的10%） ----
        risk_rate = self.get_param("风险裕量", "通用", 0.10)
        detail.risk_buffer = detail.total_direct_cost * risk_rate

        # ---- 5. 管理成本（订单金额的6%，有循环依赖需解算） ----
        mgmt_rate = self.get_param("管理费率", "通用", 0.06)
        # 管理费 = 订单金额 × 6%，订单金额 = 总成本 + 管理费 + 利润预留
        # 在保本点：订单金额 = 总成本 = 直接成本 + 风险 + 管理费
        #            = direct_total × 1.1 + 0.06 × order_amount
        # 解出：order_amount = direct_total × 1.1 / 0.94
        # 注意：保本时的管理费是基于保本价计算的
        base_before_mgmt = detail.total_direct_cost + detail.risk_buffer
        # 保本价 = base_before_mgmt / (1 - mgmt_rate)
        detail.break_even_price = base_before_mgmt / (1 - mgmt_rate)
        detail.management_cost = detail.break_even_price * mgmt_rate

        # ---- 总成本（= 保本价） ----
        detail.total_cost = base_before_mgmt + detail.management_cost
        # 理论上 total_cost == break_even_price，但保留浮点修正
        detail.break_even_price = detail.total_cost

        # ---- 目标售价（含利润预留） ----
        min_margin = self.get_param("目标毛利率", "最低", 0.25)
        max_margin = self.get_param("目标毛利率", "最高", 0.35)
        detail.target_price_25 = detail.break_even_price / (1 - min_margin)
        detail.target_price_35 = detail.break_even_price / (1 - max_margin)

        return detail

    def calculate_break_even(
        self,
        service_type: str,
        skill_level: str,
        district: str,
        duration: float,
        apartment_type: str = "3室",
        time_period: str = "工作日",
        target_margin: float = 0.30,
    ) -> Dict[str, float]:
        """
        快速计算保本价和目标售价。

        Returns:
            dict with keys: total_cost, break_even_price, target_price
        """
        detail = self.calculate_cost(
            service_type, skill_level, district,
            duration, apartment_type, time_period,
        )
        target_price = detail.break_even_price / (1 - target_margin)
        return {
            "total_cost": round(detail.total_cost, 2),
            "break_even_price": round(detail.break_even_price, 2),
            "target_price": round(target_price, 2),
            "target_margin": target_margin,
        }

    # ----------------------------------------------------------
    # 全场景批量计算
    # ----------------------------------------------------------

    def calculate_all_scenarios(
        self,
        durations: Optional[Dict[str, float]] = None,
    ) -> List[CostDetail]:
        """
        计算所有场景的保本价，覆盖 4 个维度：
        service_type × skill_level × district（必须）
        + apartment_type + time_period（可选，通过参数控制）

        Args:
            durations: 各服务类型的默认时长

        Returns:
            所有场景的 CostDetail 列表（默认 84 个核心场景）
        """
        if durations is None:
            durations = {
                "日常保洁": 3,
                "深度清洁": 4,
                "开荒保洁": 6,
                "家电清洗": 2,
            }

        results = []
        # 默认只输出核心场景（service × skill × district），共 84 个
        # 要扩展覆盖户型和时间段，用场景计算器按需调用
        for st in self.service_types:
            dur = durations.get(st, 3)
            for sl in self.skill_levels:
                for dist in self.districts:
                    detail = self.calculate_cost(st, sl, dist, dur)
                    results.append(detail)
        return results

    def calculate_full_scenarios(self) -> List[Dict]:
        """
        计算全维度场景矩阵（含户型+时间段），用于深度分析。
        共 4 × 3 × 7 × 5 × 4 = 1680 种组合。

        Returns:
            list of summary_dict
        """
        results = []
        default_durations = {"日常保洁": 3, "深度清洁": 4, "开荒保洁": 6, "家电清洗": 2}
        for st in self.service_types:
            dur = default_durations.get(st, 3)
            for sl in self.skill_levels:
                for dist in self.districts:
                    for apt in self.apartment_types:
                        for tp in self.time_periods:
                            detail = self.calculate_cost(st, sl, dist, dur, apt, tp)
                            results.append(detail.summary_dict())
        return results

    # ----------------------------------------------------------
    # 历史订单复盘
    # ----------------------------------------------------------

    def review_historical_orders(
        self, orders_csv_path: str,
    ) -> Tuple[List[Dict], List[Dict], Dict]:
        """
        用新成本模型复盘历史订单。

        Args:
            orders_csv_path: sample_orders.csv 路径

        Returns:
            (逐行复盘数据, 亏损订单明细, 亏损汇总统计)
        """
        import pandas as pd

        df = pd.read_csv(orders_csv_path)
        review_rows = []
        loss_rows = []
        total_orders = len(df)
        loss_count = 0
        total_loss_amount = 0.0

        for _, row in df.iterrows():
            st = str(row["服务类型"])
            sl = str(row["服务等级"])
            dist = str(row["城区"])
            duration = float(row["服务时长"])
            actual_price = float(row["订单金额"])

            # 用新模型计算成本
            cost = self.calculate_cost(st, sl, dist, duration)

            # 实际利润率（按新成本）
            actual_margin = (actual_price - cost.total_cost) / actual_price * 100 if actual_price > 0 else 0
            is_loss = actual_price < cost.break_even_price

            # 旧模型利润率（原有计算方式，用于对比）
            # 旧: 标准价格 = 基准单价 × 时长（含区域加价），成本 = 标准价格 × 0.85
            generic_rates = {
                ("日常保洁", "新手"): 35, ("日常保洁", "熟练"): 40, ("日常保洁", "金牌"): 45,
                ("深度清洁", "新手"): 48, ("深度清洁", "熟练"): 58, ("深度清洁", "金牌"): 68,
                ("开荒保洁", "新手"): 4, ("开荒保洁", "熟练"): 5, ("开荒保洁", "金牌"): 6,
                ("家电清洗", "新手"): 85, ("家电清洗", "熟练"): 110, ("家电清洗", "金牌"): 140,
            }
            premium_map = {"青秀": 1.15, "良庆": 1.10, "江南": 1.00, "西乡塘": 1.00, "兴宁": 1.00, "邕宁": 1.10, "武鸣": 1.25}
            base_rate = generic_rates.get((st, sl), 0)
            premium = premium_map.get(dist, 1.0)
            if st in ("日常保洁", "深度清洁"):
                old_std_price = base_rate * premium * duration
            elif st == "开荒保洁":
                prod = {"新手": 18, "熟练": 15, "金牌": 12}.get(sl, 15)
                area = duration * prod
                old_std_price = base_rate * premium * area
            else:  # 家电清洗
                units = max(1, round(duration / 1.5))
                old_std_price = base_rate * premium * units
            old_cost = old_std_price * 0.85  # 旧模型的成本
            old_margin = (actual_price - old_cost) / actual_price * 100 if actual_price > 0 else 0

            review_row = {
                "订单ID": row["订单ID"],
                "下单时间": row["下单时间"],
                "城区": dist,
                "服务类型": st,
                "服务等级": sl,
                "服务时长": duration,
                "订单金额": actual_price,
                "新模型_总成本": round(cost.total_cost, 2),
                "新模型_保本价": round(cost.break_even_price, 2),
                "新模型_实际利润率(%)": round(actual_margin, 2),
                "新模型_是否亏损": "是" if is_loss else "否",
                "新模型_亏损额": round(max(0, cost.break_even_price - actual_price), 2),
                "旧模型_估算成本": round(old_cost, 2),
                "旧模型_利润率(%)": round(old_margin, 2),
                "利润率差异(新-旧)": round(actual_margin - old_margin, 2),
                "客户评分": row.get("客户评分", ""),
            }
            review_rows.append(review_row)

            if is_loss:
                loss_count += 1
                loss_amount = max(0, cost.break_even_price - actual_price)
                total_loss_amount += loss_amount
                loss_rows.append(review_row)

        # 汇总统计
        summary = {
            "总订单数": total_orders,
            "亏损订单数": loss_count,
            "亏损率(%)": round(loss_count / total_orders * 100, 2) if total_orders > 0 else 0,
            "亏损总额(元)": round(total_loss_amount, 2),
        }

        # 按城区聚合
        dist_loss = {}
        for r in loss_rows:
            d = r["城区"]
            dist_loss.setdefault(d, {"亏损订单数": 0, "亏损总额": 0.0})
            dist_loss[d]["亏损订单数"] += 1
            dist_loss[d]["亏损总额"] += r["新模型_亏损额"]
        summary["按城区亏损"] = dist_loss

        # 按服务类型聚合
        svc_loss = {}
        for r in loss_rows:
            s = r["服务类型"]
            svc_loss.setdefault(s, {"亏损订单数": 0, "亏损总额": 0.0})
            svc_loss[s]["亏损订单数"] += 1
            svc_loss[s]["亏损总额"] += r["新模型_亏损额"]
        summary["按服务类型亏损"] = svc_loss

        # 按服务等级聚合
        sl_loss = {}
        for r in loss_rows:
            s = r["服务等级"]
            sl_loss.setdefault(s, {"亏损订单数": 0, "亏损总额": 0.0})
            sl_loss[s]["亏损订单数"] += 1
            sl_loss[s]["亏损总额"] += r["新模型_亏损额"]
        summary["按服务等级亏损"] = sl_loss

        # 亏损金额最大的 20 笔订单
        top_losses = sorted(loss_rows, key=lambda r: r["新模型_亏损额"], reverse=True)[:20]
        summary["亏损金额TOP20"] = [
            {"订单ID": r["订单ID"], "城区": r["城区"], "服务类型": r["服务类型"],
             "服务等级": r["服务等级"], "服务时长": r["服务时长"],
             "订单金额": r["订单金额"], "亏损额": r["新模型_亏损额"]}
            for r in top_losses
        ]

        # 亏损订单的服务时长分布（判断是否因超时长导致）
        from collections import Counter
        duration_buckets = Counter()
        for r in loss_rows:
            d = r["服务时长"]
            if d <= 2:
                duration_buckets["≤2h"] += 1
            elif d <= 4:
                duration_buckets["2-4h"] += 1
            elif d <= 6:
                duration_buckets["4-6h"] += 1
            elif d <= 8:
                duration_buckets["6-8h"] += 1
            else:
                duration_buckets[">8h"] += 1
        summary["亏损时长分布"] = dict(duration_buckets)

        return review_rows, loss_rows, summary

    # ----------------------------------------------------------
    # Excel 导出
    # ----------------------------------------------------------

    def export_to_excel(
        self,
        output_path: str,
        review_rows: Optional[List[Dict]] = None,
    ):
        """
        导出保本定价测算表到 Excel。

        Args:
            output_path: 输出 Excel 文件路径
            review_rows: 历史订单复盘数据（可选）
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            print("[错误] 需要安装 openpyxl: pip install openpyxl")
            return

        wb = openpyxl.Workbook()

        # ---- Sheet 1: 全场景保本价矩阵 ----
        ws1 = wb.active
        ws1.title = "全场景保本价矩阵"

        # 表头样式
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        scenarios = self.calculate_all_scenarios()
        headers = ["服务类型", "师傅等级", "城区", "服务时长(h)",
                   "人力成本", "耗材成本", "交通成本", "管理成本", "风险裕量",
                   "总成本", "保本价", "目标售价(25%毛利)", "目标售价(35%毛利)"]

        for col, h in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        for i, s in enumerate(scenarios, 2):
            sd = s.summary_dict()
            for col, h in enumerate(headers, 1):
                val = sd.get(h, "")
                cell = ws1.cell(row=i, column=col, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")

        # 列宽
        for col in range(1, len(headers) + 1):
            ws1.column_dimensions[get_column_letter(col)].width = 16

        # ---- Sheet 2: 历史订单复盘 ----
        if review_rows:
            ws2 = wb.create_sheet("历史订单复盘")
            loss_fill = PatternFill(start_color="FECACA", end_color="FECACA", fill_type="solid")

            if review_rows:
                rev_headers = list(review_rows[0].keys())
                for col, h in enumerate(rev_headers, 1):
                    cell = ws2.cell(row=1, column=col, value=h)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_align
                    cell.border = thin_border

                for i, row in enumerate(review_rows, 2):
                    for col, h in enumerate(rev_headers, 1):
                        val = row.get(h, "")
                        cell = ws2.cell(row=i, column=col, value=val)
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal="center")
                        # 亏损行标记红色
                        if row.get("新模型_是否亏损") == "是":
                            cell.fill = loss_fill

                for col in range(1, len(rev_headers) + 1):
                    ws2.column_dimensions[get_column_letter(col)].width = 18

        # ---- Sheet 3: 亏损原因汇总 ----
        if review_rows:
            ws3 = wb.create_sheet("亏损原因汇总")
            loss_rows = [r for r in review_rows if r.get("新模型_是否亏损") == "是"]
            loss_fill = PatternFill(start_color="FECACA", end_color="FECACA", fill_type="solid")
            warn_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")

            if loss_rows:
                from collections import Counter, defaultdict

                def write_table(ws, start_row, headers, data_rows, col_widths=None):
                    """写入表格，带表头样式"""
                    for col, h in enumerate(headers, 1):
                        cell = ws.cell(row=start_row, column=col, value=h)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_align
                        cell.border = thin_border
                    for i, row in enumerate(data_rows, start_row + 1):
                        for col, val in enumerate(row, 1):
                            cell = ws.cell(row=i, column=col, value=val)
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal="center")
                    if col_widths:
                        for col, w in enumerate(col_widths, 1):
                            ws.column_dimensions[get_column_letter(col)].width = w

                row = 1

                # --- 总览 ---
                ws3.cell(row=row, column=1, value="亏损总览").font = Font(bold=True, size=13)
                row += 1
                ws3.cell(row=row, column=1, value=f"总订单数: {len(review_rows)}  |  亏损订单: {len(loss_rows)}  |  亏损率: {len(loss_rows)/len(review_rows)*100:.1f}%")
                row += 2

                # --- Table 1: 按城区 ---
                ws3.cell(row=row, column=1, value="表1: 按城区亏损").font = Font(bold=True, size=12)
                row += 1
                dist_data = []
                dist_counter = Counter(r["城区"] for r in loss_rows)
                for dist, cnt in dist_counter.most_common():
                    total = sum(1 for r in review_rows if r["城区"] == dist)
                    amount = sum(r["新模型_亏损额"] for r in loss_rows if r["城区"] == dist)
                    dist_data.append([dist, cnt, total, round(cnt/total*100, 1) if total else 0, round(amount, 0)])
                write_table(ws3, row, ["城区", "亏损订单数", "总订单数", "城区亏损率(%)", "亏损总额(元)"], dist_data, [12, 12, 12, 16, 14])
                row += len(dist_data) + 3

                # --- Table 2: 按服务类型 ---
                ws3.cell(row=row, column=1, value="表2: 按服务类型亏损").font = Font(bold=True, size=12)
                row += 1
                svc_data = []
                svc_counter = Counter(r["服务类型"] for r in loss_rows)
                for svc, cnt in svc_counter.most_common():
                    total = sum(1 for r in review_rows if r["服务类型"] == svc)
                    amount = sum(r["新模型_亏损额"] for r in loss_rows if r["服务类型"] == svc)
                    svc_data.append([svc, cnt, total, round(cnt/total*100, 1) if total else 0, round(amount, 0)])
                write_table(ws3, row, ["服务类型", "亏损订单数", "总订单数", "服务亏损率(%)", "亏损总额(元)"], svc_data, [12, 12, 12, 16, 14])
                row += len(svc_data) + 3

                # --- Table 3: 按服务等级 ---
                ws3.cell(row=row, column=1, value="表3: 按服务等级亏损").font = Font(bold=True, size=12)
                row += 1
                sl_data = []
                sl_counter = Counter(r["服务等级"] for r in loss_rows)
                for sl, cnt in sl_counter.most_common():
                    total = sum(1 for r in review_rows if r["服务等级"] == sl)
                    amount = sum(r["新模型_亏损额"] for r in loss_rows if r["服务等级"] == sl)
                    sl_data.append([sl, cnt, total, round(cnt/total*100, 1) if total else 0, round(amount, 0)])
                write_table(ws3, row, ["服务等级", "亏损订单数", "总订单数", "等级亏损率(%)", "亏损总额(元)"], sl_data, [12, 12, 12, 16, 14])
                row += len(sl_data) + 3

                # --- Table 4: 亏损金额 TOP 20 ---
                ws3.cell(row=row, column=1, value="表4: 亏损金额 TOP 20（最大亏损订单）").font = Font(bold=True, size=12)
                row += 1
                top20 = sorted(loss_rows, key=lambda r: r["新模型_亏损额"], reverse=True)[:20]
                top_data = [[
                    r.get("订单ID", ""), r["城区"], r["服务类型"], r["服务等级"],
                    r["服务时长"], r["订单金额"], r["新模型_亏损额"]
                ] for r in top20]
                write_table(ws3, row,
                    ["订单ID", "城区", "服务类型", "服务等级", "服务时长(h)", "订单金额(元)", "亏损额(元)"],
                    top_data, [22, 10, 10, 10, 12, 14, 14])
                row += len(top_data) + 3

                # 亏损 TOP20 行红色标记
                for i in range(len(top_data)):
                    for col in range(1, 8):
                        ws3.cell(row=row - len(top_data) + i, column=col).fill = loss_fill

                # --- Table 5: 亏损订单时长分布 ---
                ws3.cell(row=row, column=1, value="表5: 亏损订单服务时长分布").font = Font(bold=True, size=12)
                row += 1
                dur_counter = Counter()
                for r in loss_rows:
                    d = r["服务时长"]
                    if d <= 2:
                        dur_counter["≤2h"] += 1
                    elif d <= 4:
                        dur_counter["2-4h"] += 1
                    elif d <= 6:
                        dur_counter["4-6h"] += 1
                    elif d <= 8:
                        dur_counter["6-8h"] += 1
                    else:
                        dur_counter[">8h"] += 1
                dur_data = [[bucket, cnt] for bucket, cnt in dur_counter.most_common()]
                write_table(ws3, row, ["时长区间", "亏损订单数"], dur_data, [12, 14])
                row += len(dur_data) + 3

                # --- 亏损原因分析结论 ---
                ws3.cell(row=row, column=1, value="亏损原因诊断结论").font = Font(bold=True, size=13)
                row += 1
                # 判断主要原因: 比较各维度的亏损率
                top_dist = dist_data[0]
                top_svc = svc_data[0]
                top_sl = sl_data[0]
                conclusions = [
                    f"主要亏损城区: {top_dist[0]}（亏损率 {top_dist[3]}%，亏损 {top_dist[4]:.0f} 元）",
                    f"主要亏损服务: {top_svc[0]}（亏损率 {top_svc[3]}%，亏损 {top_svc[4]:.0f} 元）",
                    f"主要亏损等级: {top_sl[0]}（亏损率 {top_sl[3]}%，亏损 {top_sl[4]:.0f} 元）",
                    f"亏损时长集中: {dur_data[0][0]}（{dur_data[0][1]} 单）",
                ]
                for c in conclusions:
                    ws3.cell(row=row, column=1, value=c).font = Font(size=11)
                    row += 1

        wb.save(output_path)
        print(f"[OK] Excel 已保存: {output_path}")

    # ----------------------------------------------------------
    # 前端计算验证
    # ----------------------------------------------------------

    def get_verification_cases(self) -> List[Dict]:
        """
        返回5个典型场景的验证用例，用于检查前端计算结果是否与Python一致。

        Returns:
            List[dict]: 每个用例包含输入参数和预期输出
        """
        cases = [
            {"service_type": "日常保洁", "skill_level": "新手", "district": "青秀", "duration": 3},
            {"service_type": "深度清洁", "skill_level": "熟练", "district": "良庆", "duration": 4},
            {"service_type": "开荒保洁", "skill_level": "金牌", "district": "武鸣", "duration": 6},
            {"service_type": "家电清洗", "skill_level": "熟练", "district": "江南", "duration": 2},
            {"service_type": "日常保洁", "skill_level": "金牌", "district": "邕宁", "duration": 5},
        ]
        results = []
        for c in cases:
            detail = self.calculate_cost(**c)
            results.append({
                "input": c,
                "output": detail.summary_dict(),
            })
        return results

    def verify_frontend(self, json_path: Optional[str] = None) -> dict:
        """
        生成前端验证用的 JSON 数据。

        Args:
            json_path: 可选，输出 JSON 文件路径

        Returns:
            dict: 验证数据
        """
        import json
        data = {
            "calculator_params": {
                "service_types": self.service_types,
                "skill_levels": self.skill_levels,
                "districts": self.districts,
                "apartment_types": self.apartment_types,
                "time_periods": self.time_periods,
            },
            "param_values": {
                cat: subs for cat, subs in self.params.items()
            },
            "verification_cases": self.get_verification_cases(),
        }
        if json_path:
            with open(json_path, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            print(f"[OK] 验证数据已保存: {json_path}")
        return data


# ============================================================
# 快捷入口
# ============================================================

def generate_cost_report():
    """生成完整的成本报告（主入口）"""
    base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    data_dir = os.path.join(base, "data", "nanning")
    output_dir = os.path.join(base, "output")
    orders_csv = os.path.join(data_dir, "sample_orders.csv")
    excel_path = os.path.join(output_dir, "南宁家政保本定价测算表.xlsx")

    print("=" * 60)
    print("  南宁家政保本定价测算 — 成本报告生成")
    print("=" * 60)

    # 1. 初始化计算器
    print("\n[1/3] 初始化成本计算器...")
    cc = CostCalculator()
    print(f"      参数加载完成: {len(cc.params)} 个参数类别")

    # 2. 全场景保本价测算
    print("\n[2/3] 计算全场景保本价...")
    scenarios = cc.calculate_all_scenarios()
    print(f"      共 {len(scenarios)} 个场景组合")

    # 3. 历史订单复盘
    print("\n[3/3] 历史订单盈利复盘...")
    if os.path.exists(orders_csv):
        review, loss_rows, summary = cc.review_historical_orders(orders_csv)
        print(f"      总订单: {summary['总订单数']}")
        print(f"      亏损订单: {summary['亏损订单数']} ({summary['亏损率(%)']}%)")
        print(f"      亏损总额: {summary['亏损总额(元)']} 元")
        print(f"      按城区: {summary.get('按城区亏损', {})}")
        print(f"      按服务类型: {summary.get('按服务类型亏损', {})}")
        cc.export_to_excel(excel_path, review)
    else:
        print(f"      [警告] 历史订单文件不存在: {orders_csv}")
        cc.export_to_excel(excel_path)

    # 4. 输出验证数据
    verify_path = os.path.join(output_dir, "calculator_verify.json")
    cc.verify_frontend(verify_path)

    print(f"\n{'=' * 60}")
    print(f"  报告生成完成!")
    print(f"  Excel: {excel_path}")
    print(f"  验证:  {verify_path}")
    print(f"{'=' * 60}")
    return cc


# ============================================================
# 测试
# ============================================================

def run_self_test():
    """运行自检：验证5个典型场景的计算结果"""
    print("=" * 60)
    print("  成本计算器自检")
    print("=" * 60)

    cc = CostCalculator()
    cases = cc.get_verification_cases()

    all_pass = True
    for i, case in enumerate(cases, 1):
        inp = case["input"]
        out = case["output"]
        print(f"\n  场景{i}: {inp['service_type']}-{inp['skill_level']}-{inp['district']} ({inp['duration']}h)")
        print(f"    总成本: {out['总成本']:.2f} 元")
        print(f"    保本价: {out['保本价']:.2f} 元")
        print(f"    目标售价(25%): {out['目标售价(25%毛利)']:.2f} 元")
        print(f"    目标售价(35%): {out['目标售价(35%毛利)']:.2f} 元")

        # 基本合理性检查
        assert out["保本价"] > 0, "保本价必须大于0"
        assert out["人力成本"] > 0, "人力成本必须大于0"
        assert out["总成本"] <= out["目标售价(25%毛利)"], "目标售价应大于总成本"
        print(f"    [OK] 合理性检查通过")

    print(f"\n{'=' * 60}")
    print(f"  自检完成: 全部 {len(cases)} 个场景通过")
    print(f"{'=' * 60}")
    return cc


if __name__ == "__main__":
    import sys
    if "--test" in sys.argv:
        run_self_test()
    else:
        generate_cost_report()
