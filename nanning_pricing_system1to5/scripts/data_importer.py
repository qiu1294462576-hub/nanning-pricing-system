#!/usr/bin/env python3
"""数据导入工具 — 支持 Excel/CSV/MySQL 导入南宁历史订单

支持格式：
- CSV文件（自动检测编码）
- Excel文件 (.xlsx)
- MySQL数据库
"""
import csv
import os
import sys
import glob
from datetime import datetime

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(BASE, "data/nanning")
SAMPLE_DIR = os.path.join(BASE, "data/imports")


def detect_encoding(filepath):
    """检测文件编码"""
    import chardet
    with open(filepath, "rb") as f:
        raw = f.read(10000)
        result = chardet.detect(raw)
        return result.get("encoding", "utf-8")


def import_csv(filepath):
    """导入CSV文件"""
    enc = detect_encoding(filepath)
    print(f"  编码检测: {enc}")
    rows = []
    with open(filepath, "r", encoding=enc, errors="replace") as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append(row)
    print(f"  读取 {len(rows)} 条记录")
    return rows


def import_excel(filepath):
    """导入Excel文件"""
    try:
        import openpyxl
    except ImportError:
        print("  [错误] 需要安装 openpyxl: pip install openpyxl")
        return []
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = {}
        for i, val in enumerate(row):
            if i < len(headers) and headers[i]:
                row_dict[headers[i]] = val
        if row_dict:
            rows.append(row_dict)
    print(f"  读取 {len(rows)} 条记录")
    return rows


def import_mysql(host="localhost", port=3306, user="root", password="", database="nanning_pricing"):
    """从MySQL导入订单数据"""
    try:
        import pymysql
    except ImportError:
        print("  [错误] 需要安装 pymysql: pip install pymysql")
        return []
    try:
        conn = pymysql.connect(
            host=host, port=port, user=user,
            password=password, database=database,
            charset="utf8mb4"
        )
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM orders")
        columns = [desc[0] for desc in cursor.description]
        rows = [dict(zip(columns, row)) for row in cursor.fetchall()]
        print(f"  从MySQL读取 {len(rows)} 条记录")
        conn.close()
        return rows
    except Exception as e:
        print(f"  [错误] MySQL连接失败: {e}")
        return []


def standardize_fields(rows):
    """字段映射：将不同来源的字段名映射为标准字段"""
    field_map = {
        "order_id": "订单ID",
        "order_no": "订单ID",
        "create_time": "下单时间",
        "order_time": "下单时间",
        "book_time": "下单时间",
        "city": "城区",
        "region": "城区",
        "district": "城区",
        "service_type": "服务类型",
        "service": "服务类型",
        "service_level": "服务等级",
        "level": "服务等级",
        "duration": "服务时长",
        "hours": "服务时长",
        "total_amount": "订单金额",
        "price": "订单金额",
        "amount": "订单金额",
    }
    std_rows = []
    for row in rows:
        new_row = {}
        for old_k, v in row.items():
            k = old_k.strip()
            if k in field_map:
                new_row[field_map[k]] = v
            else:
                new_row[k] = v
        std_rows.append(new_row)
    print(f"  字段标准化完成")
    return std_rows


def save_to_nanning(rows):
    """保存到标准格式CSV"""
    os.makedirs(DATA_DIR, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out = os.path.join(DATA_DIR, f"imported_orders_{ts}.csv")
    with open(out, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=rows[0].keys())
        w.writeheader()
        w.writerows(rows)
    print(f"  已保存: {out}")
    return out


def main():
    print("=" * 50)
    print("数据导入工具 — 南宁家政定价系统")
    print("=" * 50)

    if len(sys.argv) < 2:
        print("\n用法:")
        print("  python scripts/data_importer.py <文件路径>")
        print("  python scripts/data_importer.py --csv <路径>")
        print("  python scripts/data_importer.py --excel <路径>")
        print("  python scripts/data_importer.py --mysql")
        print("\n支持格式: .csv, .xlsx")
        print("示例: python scripts/data_importer.py data/imports/orders.xlsx")
        return

    arg = sys.argv[1]
    rows = []

    if arg == "--mysql":
        rows = import_mysql()
    elif arg == "--csv" and len(sys.argv) > 2:
        rows = import_csv(sys.argv[2])
    elif arg == "--excel" and len(sys.argv) > 2:
        rows = import_excel(sys.argv[2])
    elif arg.endswith(".csv"):
        rows = import_csv(arg)
    elif arg.endswith(".xlsx"):
        rows = import_excel(arg)
    else:
        print(f"[错误] 不支持的文件格式: {arg}")
        return

    if rows:
        rows = standardize_fields(rows)
        out = save_to_nanning(rows)
        print(f"\n完成! {len(rows)} 条记录已导入")
    else:
        print("\n未读取到数据")


if __name__ == "__main__":
    main()
